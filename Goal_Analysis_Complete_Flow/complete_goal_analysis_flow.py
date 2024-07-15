from openpyxl import load_workbook
import psycopg2
from openpyxl.styles import PatternFill,Font
import shutil



def map_hr_goals(goal,sheet_name):        
    file_path = 'GoalanAlysisReport.xlsx'  # Replace with your actual file path
    workbook = load_workbook(file_path)
    sheet = workbook.worksheets[0]

    # Specify the row and column you want to read
    found_goal = False
    found_audio = False
    required_row = 0
    required_col = 0
    goal = str(goal).lower()
    sheet_name = str(sheet_name).lower()
    
    for row_num in range(3,sheet.max_row):
        if not found_goal:
            for col_num in range(2, sheet.max_column + 1):
                cell_value = sheet.cell(row=row_num, column=col_num).value
                cell_value = str(cell_value).lower()
                print(cell_value)
                if cell_value == goal :
                    print(f"Found {goal} at column {col_num}")
                    found_goal = True
                    required_col = col_num
                    break
                
        if not found_audio:
            cell_value = sheet.cell(row=row_num, column=1).value
            cell_value = str(cell_value).lower()
            if cell_value == sheet_name :
                print(f"Found {sheet_name} at row {row_num}")
                found_audio = True
                required_row = row_num

        # Stop further iterations if both conditions are met
        if found_goal and found_audio:
            print(f"required row and column: {required_row}, {required_col}")
            sheet.cell(row=required_row, column=required_col).value = "✅"
            print(f"Added a checkmark (✅) in cell {required_row}:{required_col}")
            workbook.save('GoalanAlysisReport.xlsx')
            break
        
def fetch_goal_information(task_id, sheet_name):
    conn = psycopg2.connect(
    host='qa-psql.postgres.database.azure.com',
    database='postgres',
    user='v3qa_postgres',
    password='Y&F8zn.s+!r3BK%TJWMq<$'
)
    

# Create a cursor object
    cursor = conn.cursor()

# Execute a query
    cursor.execute(
            'SELECT DISTINCT (g.name) FROM goals_analysis ga '
            'LEFT JOIN goals_snippet_mapping gsm ON gsm.goal_analysis_id = ga.id '
            'LEFT JOIN goals g ON g.id = ga.goal_id '
            'WHERE ga.task_id = %s;', (task_id,)
        )

    # Fetch all results
    results = cursor.fetchall()
        # Assuming the result is a list of goal names in the variable 'goals'
        
    goals = [result[0] for result in results]
    # Example goal list, replace with actual query results

    
    for goal in goals:
        print(goal)
        map_v3_goal_in_sheet(goal, sheet_name)
        
def map_v3_goal_in_sheet(goal, sheet_name):
    workbook = load_workbook('GoalanAlysisReport.xlsx')
    sheet = workbook.worksheets[1]

    found_goal = False
    found_audio = False
    required_row = 0
    required_col = 0
    goal = str(goal).lower()
    sheet_name = str(sheet_name).lower()
    for row_num in range(3, sheet.max_row+1):
        if not found_goal:
            for col_num in range(2, sheet.max_column + 1):
                cell_value = sheet.cell(row=row_num, column=col_num).value
                cell_value = str(cell_value).lower()
                print(cell_value)
                if cell_value == goal :
                    print(f"Found {goal} at column {col_num}")
                    found_goal = True
                    required_col = col_num
                    break
                
        if not found_audio:
            cell_value = sheet.cell(row=row_num, column=1).value
            cell_value = str(cell_value).lower()
            if cell_value == sheet_name :
                print(f"Found {sheet_name} at row {row_num}")
                found_audio = True
                required_row = row_num

        # Stop further iterations if both conditions are met
        if found_goal and found_audio:
            print(f"required row and column: {required_row}, {required_col}")
            sheet.cell(row=required_row, column=required_col).value = "✅"
            print(f"Added a checkmark (✅) in cell {required_row}:{required_col}")
            workbook.save('GoalanAlysisReport.xlsx')
            break    
        
def findTP(HR_goal_audio_map,V3_goal_audio_map,TotalTP):
    TP_of_particular_goal = {}
    file_path = 'GoalanAlysisReport.xlsx'
    workbook = load_workbook(file_path)
    sheet = workbook.worksheets[2]
        
    column = 2
    for goal, hr_audio_files in HR_goal_audio_map.items():
        TP = 0
        if goal in V3_goal_audio_map:
            v3_audio_files = V3_goal_audio_map[goal]
            
            # Check if the audio file names in HR_goal_audio_map are also in V3_goal_audio_map for the same goal
            for hr_audio in hr_audio_files:
                if hr_audio in v3_audio_files:
                    TP = TP+1
                    TotalTP = TotalTP+1
        print(f"TP of {goal}: {TP}")
        sheet.cell(row=4, column=column, value=TP) 
        sheet.cell(row=4, column=column).font = Font(color="000000")
        column = column+1
    #       Save the workbook
        workbook.save(file_path)
        TP_of_particular_goal[goal] = TP
        
    return TP_of_particular_goal,TotalTP

def findTN(HR_goal_not_found_audio_map,V3_goal_not_found_audio_map,TotalTN):
    
    TN_of_particular_goal = {}
    file_path = 'GoalanAlysisReport.xlsx'
    workbook = load_workbook(file_path)
    sheet = workbook.worksheets[2]
    
    column1 = 2
    
    for goal, hr_audio_files in HR_goal_not_found_audio_map.items():
        TN = 0
    # Check if the goal exists in V3_goal_audio_map
        if goal in V3_goal_not_found_audio_map:
            v3_audio_files = V3_goal_not_found_audio_map[goal]
        
        # Check if the audio file names in HR_goal_audio_map are also in V3_goal_audio_map for the same goal
            for hr_audio in hr_audio_files:
                if hr_audio in v3_audio_files:
                    TN = TN+1
                    TotalTN = TotalTN + 1
        print(f"TN of {goal}: {TN}")
        sheet.cell(row=5, column=column1, value=TN)
        sheet.cell(row=5, column=column1).font = Font(color="000000")
        column1 = column1+1
#        Save the workbook
        workbook.save(file_path)
        TN_of_particular_goal[goal] = TN
        
    return TN_of_particular_goal,TotalTN
        
        
def findFP(HR_goal_not_found_audio_map, V3_goal_found_audio_map,TotalFP):
    FP_of_particular_goal = {}
    file_path = 'GoalanAlysisReport.xlsx'
    workbook = load_workbook(file_path)
    sheet = workbook.worksheets[2]
    
    column1 = 2
    
    for goal, v3_audio_files in V3_goal_found_audio_map.items():
        FP = 0
    # Check if the goal exists in V3_goal_audio_map
        if goal in HR_goal_not_found_audio_map:
            HR_audio_files = HR_goal_not_found_audio_map[goal]
        
        # Check if the audio file names in HR_goal_audio_map are also in V3_goal_audio_map for the same goal
            for V3_audio in v3_audio_files:
                if V3_audio in HR_audio_files:
                    FP = FP+1
                    TotalFP = TotalFP+1
        print(f"FP of {goal}: {FP}")
        FP_of_particular_goal[goal] = FP
        sheet.cell(row=6, column=column1, value=FP)
        sheet.cell(row=6, column=column1).font = Font(color="000000")
        column1 = column1+1
#        Save the workbook
        workbook.save(file_path)
    return FP_of_particular_goal,TotalFP
        
        
def findFN(HR_goal_found_audio_map,V3_goal_not_found_audio_map,TotalFN):
    FN_of_particular_goal = {}
    file_path = 'GoalanAlysisReport.xlsx'
    workbook = load_workbook(file_path)
    sheet = workbook.worksheets[2]
    
    column1 = 2
    
    for goal, v3_audio_files in V3_goal_not_found_audio_map.items():
        FN = 0
    # Check if the goal exists in V3_goal_audio_map
        if goal in HR_goal_found_audio_map:
            HR_audio_files = HR_goal_found_audio_map[goal]
        
        # Check if the audio file names in HR_goal_audio_map are also in V3_goal_audio_map for the same goal
            for V3_audio in v3_audio_files:
                if V3_audio in HR_audio_files:
                    FN = FN+1
                    TotalFN = TotalFN+1
        print(f"FN of {goal}: {FN}")
        FN_of_particular_goal[goal] = FN
        sheet.cell(row=7, column=column1, value=FN)
        sheet.cell(row=7, column=column1).font = Font(color="000000")
        column1 = column1+1
#        Save the workbook
        workbook.save(file_path)
        
    return FN_of_particular_goal,TotalFN
        
def insert_accuracy(accuracy,goal,col2):
    file_path = 'GoalanAlysisReport.xlsx'
    workbook = load_workbook(file_path)
    sheet = workbook.worksheets[2]
        
    sheet.cell(row=8, column=col2, value=accuracy)
    if(accuracy >= 75):
        green_fill = PatternFill(start_color="CAFAAD", end_color="CAFAAD", fill_type="solid")
        sheet.cell(row=8, column=col2).fill = green_fill
        sheet.cell(row=8, column=col2).font = Font(color="000000")
    else:
        red_fill = PatternFill(start_color="FAB7AD", end_color="FAB7AD", fill_type="solid")
        sheet.cell(row=8, column=col2).fill = red_fill
        sheet.cell(row=8, column=col2).font = Font(color="000000")
        
    workbook.save(file_path)
    print(f"Accuracy of goal {goal} inserted")
    return file_path



def insert_recall(recall,goal,col4):
    file_path = 'GoalanAlysisReport.xlsx'
    workbook = load_workbook(file_path)
    sheet = workbook.worksheets[2]
        
    sheet.cell(row=9, column=col4, value=recall)
    if(recall >= 75):
        green_fill = PatternFill(start_color="CAFAAD", end_color="CAFAAD", fill_type="solid")
        sheet.cell(row=9, column=col4).fill = green_fill
        sheet.cell(row=9, column=col4).font = Font(color="000000")
    else:
        red_fill = PatternFill(start_color="FAB7AD", end_color="FAB7AD", fill_type="solid")
        sheet.cell(row=9, column=col4).fill = red_fill
        sheet.cell(row=9, column=col4).font = Font(color="000000")
        
    workbook.save(file_path)
    print(f"Precision of goal {goal} inserted")
    
    return file_path   

def insert_precision(precision,goal,col3):
    file_path = 'GoalanAlysisReport.xlsx'
    workbook = load_workbook(file_path)
    sheet = workbook.worksheets[2]
        
    sheet.cell(row=10, column=col3, value=precision)
    if(precision >= 75):
        green_fill = PatternFill(start_color="CAFAAD", end_color="CAFAAD", fill_type="solid")
        sheet.cell(row=10, column=col3).fill = green_fill
        sheet.cell(row=10, column=col3).font = Font(color="000000")
    else:
        red_fill = PatternFill(start_color="FAB7AD", end_color="FAB7AD", fill_type="solid")
        sheet.cell(row=10, column=col3).fill = red_fill
        sheet.cell(row=10, column=col3).font = Font(color="000000")
        
    workbook.save(file_path)
    print(f"Precision of goal {goal} inserted")
    
    return file_path  

def add_sheetname_in_HR_Mapping(sheet_names):
    file_path = 'GoalanAlysisReport.xlsx'  
    workbook = load_workbook(file_path)
    sheet = workbook.worksheets[0]
    total_audio_files = len(sheet_names)
    sheet_index = 0
    for row in sheet.iter_rows(min_row=4, max_row=total_audio_files+4):
        row[0].value = sheet_names[sheet_index]
        sheet_index = sheet_index + 1
        if(sheet_index == len(sheet_names)):
            break
            
    workbook.save(file_path)
    
def add_sheetname_in_V3_Mapping(sheet_names):
    file_path = 'GoalanAlysisReport.xlsx'  
    workbook = load_workbook(file_path)
    sheet = workbook.worksheets[1]
    total_audio_files = len(sheet_names)
    sheet_index = 0
    for row in sheet.iter_rows(min_row=4,max_row=total_audio_files+4):
        row[0].value = sheet_names[sheet_index]
        sheet_index = sheet_index + 1
        if(sheet_index == len(sheet_names)):
            break
            
    workbook.save(file_path)

if __name__ == '__main__':
    
    # Making the copy of GoalanAlysisReport Excel File
    
    original_file = 'goalanalysis_template.xlsx'

# Define the path to the new Excel file
    copy_file = 'GoalanAlysisReport.xlsx'

# Use shutil to copy the file
    shutil.copyfile(original_file, copy_file)

    file_path = 'snippets.xlsx'  
    workbook = load_workbook(file_path)
    sheet_names = workbook.sheetnames
    TotalTP = 0
    TotalTN = 0
    TotalFP = 0
    TotalFN = 0
    TotalAccuracy = 0
    TotalPrecision = 0
    TotoalRecall = 0 

    Goals_In_Each_Sheet = {}
    
    # Adding the sheet name in the HR goal mapping sheet as column
    add_sheetname_in_HR_Mapping(sheet_names)
    
    
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        print(f"Processing sheet: {sheet_name}")

        # Identify the merged cell for "goals mapping" (assuming it's in the third row)
        merged_cells = sheet.merged_cells.ranges
        goals_mapping_range = None
        for merged_cell in merged_cells:
            if '3' in str(merged_cell):
                goals_mapping_range = merged_cell
                break

        # Get the start and end columns of the merged cell
        if goals_mapping_range:
            start_col, end_col = goals_mapping_range.min_col, goals_mapping_range.max_col

            # Collect unique values from columns under "goals mapping"
            unique_values = set()
            for col in range(start_col, end_col + 1):
                for row in range(4, sheet.max_row + 1):  # Assuming data starts from row 4
                    cell_value = sheet.cell(row=row, column=col).value
                    # print(cell_value)
                    if cell_value is not None and cell_value != 'None':
                        if cell_value == "Usecases":
                            cell_value = "Benefits and Features"
                        unique_values.add(cell_value)
                        unique_goals = list(unique_values)

            print(f"Unique values in the 'goals mapping' columns: for sheet {sheet_name}")
            print(unique_goals)
            Goals_In_Each_Sheet[sheet_name] = unique_goals
        else:
            print("No merged cell found for 'goals mapping' in the third row.")
           
    
    for key, value in Goals_In_Each_Sheet.items():
        print(f"Key: {key}, Value: {value}")
        for goal in value:
            map_hr_goals(goal,key)
            
       
       
    # Map V3 detected Goals         
    file_path = 'snippets.xlsx'  
    workbook = load_workbook(file_path)
    sheet_names = workbook.sheetnames
    
    # Adding sheet name as the column in the v3 mapping sheet
    add_sheetname_in_V3_Mapping(sheet_names)


    if len(sheet_names) < 1:
        print("The workbook does not contain at least 1 sheets.")
    else:
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            print(f"Processing sheet: {sheet_name}")

            cell_value = sheet.cell(row=2, column=1).value

            if cell_value:
                task_id = cell_value.split()[-1]
                print(task_id)
                fetch_goal_information(task_id, sheet_name)



    # Making Goal Analysis report
    file_path = 'GoalanAlysisReport.xlsx'
    workbook = load_workbook(file_path)
    sheet = workbook.worksheets[0]
    
    max_row = len(sheet_names)

    # Extract goal names from the third row (index 2, considering 0-indexed for openpyxl)
    HR_goal_row_index = 3
    HR_goal_found_audio_map = {}
    HR_goal_not_found_audio_map = {}
    # Iterate over all columns starting from the 2nd column
    for col in range(2, sheet.max_column + 1):
        goal_name = sheet.cell(row=HR_goal_row_index, column=col).value
        
        print (f"goal_name : {goal_name}")
        # Initialize the list for storing audio files for the current goal
        
        
        HR_audio_files = []
        HR_audio_not_found_file = []

        # Iterate over the column and collect audio file names where "✅" is present
        for row in range(4, max_row+4):
            
            if sheet.cell(row=row, column=col).value == '✅':
                audio_file_name = sheet.cell(row=row, column=1).value
                print(f"audio_file_name : {audio_file_name}")
                HR_audio_files.append(audio_file_name)
            else:
                audio_file_name = sheet.cell(row=row, column=1).value
                HR_audio_not_found_file.append(audio_file_name)
        
        # Store the goal and its corresponding audio files in the dictionary
        HR_goal_found_audio_map[goal_name] = HR_audio_files # Remove duplicates
        HR_goal_not_found_audio_map[goal_name] = HR_audio_not_found_file
    
    
    
    file_path = 'GoalanAlysisReport.xlsx'
    workbook = load_workbook(file_path)
    sheet = workbook.worksheets[1]    
        
    V3_goal_found_audio_map = {}
    V3_goal_not_found_audio_map = {}
    V3_goal_row_index = 3
    # Iterate over all columns starting from the 2nd column
    for col in range(2, sheet.max_column + 1):
        goal_name = sheet.cell(row=V3_goal_row_index, column=col).value
        
        print (f"goal_name : {goal_name}")
        # Initialize the list for storing audio files for the current goal
        V3_audio_files = []
        V3_audio_not_found_file = []
        

        # Iterate over the column and collect audio file names where "✅" is present
        for row in range(4, max_row+4):
            
            if sheet.cell(row=row, column=col).value == '✅':
                audio_file_name = sheet.cell(row=row, column=1).value
                # print(f"audio_file_name : {audio_file_name}")
                V3_audio_files.append(audio_file_name)
            else:
                audio_file_name = sheet.cell(row=row, column=1).value
                V3_audio_not_found_file.append(audio_file_name)
        
        # Store the goal and its corresponding audio files in the dictionary
        V3_goal_found_audio_map[goal_name] = V3_audio_files  # Remove duplicates
        V3_goal_not_found_audio_map[goal_name] = V3_audio_not_found_file
        
    for key in HR_goal_found_audio_map:
        value = HR_goal_found_audio_map[key]
        print(f"HR_Goal: {key}, HR_Audio: {value}")


    for key in V3_goal_found_audio_map:
        value = V3_goal_found_audio_map[key]
        print(f"V3_Goal: {key}, V3_Audio: {value}")
        
    for key in HR_goal_not_found_audio_map:
        value = HR_goal_not_found_audio_map[key]
        print(f"HR_Goal: {key}, HR_Audio: {value}")


    for key in V3_goal_not_found_audio_map:
        value = V3_goal_not_found_audio_map[key]
        print(f"V3_Goal: {key}, V3_Audio: {value}")
        
        
    TP_of_particular_goal,TotalTP = findTP(HR_goal_found_audio_map,V3_goal_found_audio_map,TotalTP)    

    TN_of_particular_goal,TotalTN = findTN(HR_goal_not_found_audio_map, V3_goal_not_found_audio_map,TotalTN)

    FP_of_particular_goal,TotalFP = findFP(HR_goal_not_found_audio_map, V3_goal_found_audio_map,TotalFP)

    FN_of_particular_goal,TotalFN = findFN(HR_goal_found_audio_map,V3_goal_not_found_audio_map,TotalFN)

    col2 = 2
    for key in TP_of_particular_goal:
        num = TP_of_particular_goal[key] + TN_of_particular_goal[key]
        denom = TP_of_particular_goal[key] + TN_of_particular_goal[key] + FP_of_particular_goal[key] + FN_of_particular_goal[key]
        try:
            accuracy = num / denom * 100
        except Exception as e:
            print(f"An unexpected error occurred: {e}")
            accuracy = 0
        accuracy = round(accuracy,2)
        insert_accuracy(accuracy,key,col2) 
        print(f"accuracy of goal : {key} :- {accuracy} %")
        col2 = col2+1

    col3 = 2
    for key in TP_of_particular_goal:
        num = TP_of_particular_goal[key] 
        denom = TP_of_particular_goal[key] + FP_of_particular_goal[key]
        try:
            precision = num / denom * 100
        except Exception as e:
            print(f"An unexpected error occurred: {e}")
            precision = 0
        precision = round(precision,2)
        insert_precision(precision,key,col3) 
        print(f"Precision of goal : {key} :- {precision} %")
        col3 = col3+1


    col4 = 2
    for key in TP_of_particular_goal:
        num = TP_of_particular_goal[key] 
        denom = TP_of_particular_goal[key] + FN_of_particular_goal[key] 
        try:
            recall = num / denom * 100
        except Exception as e:
            print(f"An unexpected error occurred: {e}")
            recall = 0
        recall = round(recall,2)
        insert_recall(recall,key,col4) 
        print(f"Recall of goal : {key} :- {recall} %")
        col4 = col4+1


    # 
    
    file_path = 'GoalanAlysisReport.xlsx'
    workbook = load_workbook(file_path)
    sheet = workbook.worksheets[2]
    
    accuracy_num = TotalTP + TotalTN
    accuracy_deno = TotalTP + TotalTN + TotalFP + TotalFN
    
    try:
        TotalAccuracy = accuracy_num / accuracy_deno * 100
    except Exception as e:
            print(f"An unexpected error occurred: {e}")
            TotalAccuracy = 0
            
    TotalAccuracy = round(TotalAccuracy,2)
    print(f"TotalAccuracy : {TotalAccuracy}")
    
    sheet.cell(row=12, column=1, value="Total Accuracy")
    sheet.cell(row=12, column=2, value= TotalAccuracy)
    if(TotalAccuracy >= 75):
        green_fill = PatternFill(start_color="CAFAAD", end_color="CAFAAD", fill_type="solid")
        sheet.cell(row=12, column=2).fill = green_fill
        sheet.cell(row=12, column=2).font = Font(color="000000")
    else:
        red_fill = PatternFill(start_color="FAB7AD", end_color="FAB7AD", fill_type="solid")
        sheet.cell(row=12, column=2).fill = red_fill
        sheet.cell(row=12, column=2).font = Font(color="000000")

    precision_num = TotalTP
    precision_deno = TotalTP + TotalFP
    
    try:
        TotalPrecision = precision_num / precision_deno * 100
    except Exception as e:
            print(f"An unexpected error occurred: {e}")
            TotalPrecision = 0
    TotalPrecision = round(TotalPrecision,2)   
    
    print(f"TotalPrecision : {TotalPrecision}")    
    
    sheet.cell(row=13, column=1, value="Total Precision")
    sheet.cell(row=13, column=2, value= TotalPrecision)
    
    if(TotalPrecision >= 75):
        green_fill = PatternFill(start_color="CAFAAD", end_color="CAFAAD", fill_type="solid")
        sheet.cell(row=13, column=2).fill = green_fill
        sheet.cell(row=13, column=2).font = Font(color="000000")
    else:
        red_fill = PatternFill(start_color="FAB7AD", end_color="FAB7AD", fill_type="solid")
        sheet.cell(row=13, column=2).fill = red_fill
        sheet.cell(row=13, column=2).font = Font(color="000000")
    
    
    recall_num = TotalTP
    recall_deno = TotalTP + TotalFP
    
    try:
        TotalRecall = recall_num / recall_deno * 100
    except Exception as e:
            print(f"An unexpected error occurred: {e}")
            TotalRecall = 0
    TotalRecall = round(TotalRecall,2)       
    print(f"TotalRecall{TotalRecall}")
    
    sheet.cell(row=14, column=1, value="Total Recall")
    sheet.cell(row=14, column=2, value= TotalRecall)
    
    if(TotalPrecision >= 75):
        green_fill = PatternFill(start_color="CAFAAD", end_color="CAFAAD", fill_type="solid")
        sheet.cell(row=14, column=2).fill = green_fill
        sheet.cell(row=14, column=2).font = Font(color="000000")
    else:
        red_fill = PatternFill(start_color="FAB7AD", end_color="FAB7AD", fill_type="solid")
        sheet.cell(row=14, column=2).fill = red_fill
        sheet.cell(row=14, column=2).font = Font(color="000000")
    
            
    workbook.save(file_path)